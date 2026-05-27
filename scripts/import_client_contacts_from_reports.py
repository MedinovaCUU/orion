#!/usr/bin/env python3
from __future__ import annotations

import argparse
import difflib
import re
import subprocess
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


DB_CONTAINER = "supabase_db_Biosystems_Project"
DB_NAME = "postgres"
DB_USER = "postgres"

INVALID_CONTACTS = {
    "",
    ".",
    "-",
    "N/A",
    "NA",
    "N/D",
    "ND",
    "NO APLICA",
    "NOAPLICA",
    "SIN DATO",
    "SIN CONTACTO",
    "NINGUNO",
    "0",
    "1",
    "Q.",
    "Q",
}

COMMON_CONTACT_PREFIXES = re.compile(
    r"\b("
    r"Q|Q\.|QFB|Q\.F\.B|QBP|Q\.B\.P|QBP\.|QFB\.|QBB|Q\.B\.B|QUIMICA|QUIMICO|QUIMICOA|"
    r"LIC|LIC\.|ING|ING\.|DR|DR\.|DRA|DRA\.|MVZ|MVZ\.|MTRA|MTRA\.|MTRO|MTRO\.|"
    r"TSU|TLC|T\.L\.C|TLC\.|BIOLOGO|BIOLOGA|BIOL|BIOL\."
    r")\b\.?",
    re.IGNORECASE,
)


@dataclass
class ReportRow:
    serial_norm: str
    cliente: str
    contacto: str
    contacto_key: str
    telefono: str
    telefono_key: str
    fecha_servicio: date | None


@dataclass
class EquipmentRow:
    id: str
    numero_serie: str
    serial_norm: str
    cliente_id: int | None
    cliente_nombre: str
    current_rank: int


@dataclass
class SignalAggregate:
    count: int = 0
    weight: float = 0.0
    latest: date = date.min
    variants: Counter[str] = field(default_factory=Counter)

    def add(self, variant: str, weight: float, when: date | None) -> None:
        self.count += 1
        self.weight += weight
        if when and when > self.latest:
            self.latest = when
        if variant:
            self.variants[variant] += 1


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"None", "nan"}:
        return ""
    return re.sub(r"\s+", " ", text)


def normalize_serial(value: object) -> str:
    digits = re.sub(r"\D+", "", clean_text(value))
    digits = digits.lstrip("0")
    return digits or "0"


def strip_accents(value: str) -> str:
    return "".join(
        character for character in unicodedata.normalize("NFD", value) if unicodedata.category(character) != "Mn"
    )


def normalize_name(value: object) -> str:
    text = strip_accents(clean_text(value))
    return re.sub(r"[^A-Z0-9]+", "", text.upper())


def prettify_contact(value: str) -> str:
    text = clean_text(value)
    if not text:
        return ""
    if text.upper() == text and any(character.isalpha() for character in text):
        text = " ".join(word.capitalize() for word in text.lower().split())
    return text


def normalize_contact_key(value: object) -> str:
    text = strip_accents(clean_text(value)).upper()
    text = COMMON_CONTACT_PREFIXES.sub(" ", text)
    text = re.sub(r"[^A-Z0-9 ]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    if text in INVALID_CONTACTS or len(text) < 4:
        return ""
    return text


def clean_contact_display(value: object) -> str:
    text = clean_text(value)
    upper = strip_accents(text).upper()
    if upper in INVALID_CONTACTS:
        return ""
    return prettify_contact(text)


def normalize_phone_key(value: object) -> str:
    digits = re.sub(r"\D+", "", clean_text(value))
    if digits.startswith("52") and len(digits) >= 12:
        digits = digits[-10:]
    if len(digits) < 7:
        return ""
    if set(digits) == {digits[0]}:
        return ""
    if digits in {"0", "1"}:
        return ""
    return digits


def clean_phone_display(value: object) -> str:
    digits = normalize_phone_key(value)
    return digits


def parse_service_date(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = clean_text(value)
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%b-%y", "%d-%b-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def similarity_score(left: str, right: str) -> float:
    left_norm = normalize_name(left)
    right_norm = normalize_name(right)
    if not left_norm or not right_norm:
        return 0.0
    if left_norm == right_norm:
        return 1.0
    if left_norm in right_norm or right_norm in left_norm:
        return 0.92
    return difflib.SequenceMatcher(None, left_norm, right_norm).ratio()


def signal_sort_key(item: tuple[str, SignalAggregate]) -> tuple[float, int, int, int, str]:
    value, aggregate = item
    latest_ordinal = aggregate.latest.toordinal() if aggregate.latest != date.min else 0
    variant_quality = max((len(variant) for variant in aggregate.variants), default=0)
    return (aggregate.weight, aggregate.count, latest_ordinal, variant_quality, value)


def choose_variant(aggregate: SignalAggregate) -> str:
    if not aggregate.variants:
        return ""
    best_variant, _ = sorted(
        aggregate.variants.items(),
        key=lambda item: (-item[1], -len(item[0]), item[0]),
    )[0]
    return best_variant


def sql_literal(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


def run_psql(sql: str) -> str:
    result = subprocess.run(
        [
            "docker",
            "exec",
            DB_CONTAINER,
            "psql",
            "-U",
            DB_USER,
            "-d",
            DB_NAME,
            "-At",
            "-F",
            "\t",
            "-c",
            sql,
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    return result.stdout


def execute_sql(sql: str) -> None:
    subprocess.run(
        [
            "docker",
            "exec",
            "-i",
            DB_CONTAINER,
            "psql",
            "-U",
            DB_USER,
            "-d",
            DB_NAME,
            "-v",
            "ON_ERROR_STOP=1",
        ],
        check=True,
        input=sql,
        text=True,
    )


def load_report_rows(workbook_path: Path) -> list[ReportRow]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    headers = [clean_text(cell) for cell in next(rows)]
    indexes = {header: index for index, header in enumerate(headers) if header}

    required = ["Cliente", "Persona de contacto", "Telefono", "Numero de Serie", "Fecha Servicio"]
    missing = [header for header in required if header not in indexes]
    if missing:
        raise RuntimeError(f"faltan_columnas_requeridas: {', '.join(missing)}")

    report_rows: list[ReportRow] = []
    for row in rows:
        serial_norm = normalize_serial(row[indexes["Numero de Serie"]])
        if not serial_norm or serial_norm == "0":
            continue

        contacto = clean_contact_display(row[indexes["Persona de contacto"]])
        telefono = clean_phone_display(row[indexes["Telefono"]])
        contacto_key = normalize_contact_key(contacto)
        telefono_key = normalize_phone_key(telefono)

        if not contacto_key and not telefono_key:
            continue

        report_rows.append(
            ReportRow(
                serial_norm=serial_norm,
                cliente=clean_text(row[indexes["Cliente"]]),
                contacto=contacto,
                contacto_key=contacto_key,
                telefono=telefono,
                telefono_key=telefono_key,
                fecha_servicio=parse_service_date(row[indexes["Fecha Servicio"]]),
            )
        )

    return report_rows


def load_equipment_rows() -> dict[str, list[EquipmentRow]]:
    sql = """
    WITH ranked AS (
      SELECT
        equipos.id,
        equipos.numero_serie,
        equipos.cliente_id,
        COALESCE(clientes.razon_social, '') AS cliente_nombre,
        ROW_NUMBER() OVER (
          PARTITION BY equipos.numero_serie
          ORDER BY
            (equipos.fecha_fin IS NULL) DESC,
            equipos.fecha_fin DESC NULLS LAST,
            equipos.fecha_inicio DESC NULLS LAST,
            equipos.creado_en DESC,
            equipos.id DESC
        ) AS rn
      FROM public.equipos
      LEFT JOIN public.clientes ON clientes.id = equipos.cliente_id
    )
    SELECT id, numero_serie, COALESCE(cliente_id::text, ''), cliente_nombre, rn
    FROM ranked
    ORDER BY numero_serie, rn;
    """

    equipment_by_serial: dict[str, list[EquipmentRow]] = defaultdict(list)
    for row in run_psql(sql).splitlines():
        if not row.strip():
            continue
        equipment_id, numero_serie, cliente_id_text, cliente_nombre, current_rank = row.split("\t")
        serial_norm = normalize_serial(numero_serie)
        equipment_by_serial[serial_norm].append(
            EquipmentRow(
                id=equipment_id,
                numero_serie=numero_serie,
                serial_norm=serial_norm,
                cliente_id=int(cliente_id_text) if cliente_id_text else None,
                cliente_nombre=cliente_nombre,
                current_rank=int(current_rank),
            )
        )
    return equipment_by_serial


def choose_equipment(report_row: ReportRow, candidates: list[EquipmentRow]) -> tuple[EquipmentRow | None, float, bool]:
    viable = [candidate for candidate in candidates if candidate.cliente_id is not None]
    if not viable:
        return None, 0.0, False
    viable = sorted(viable, key=lambda candidate: candidate.current_rank)
    if len({candidate.cliente_id for candidate in viable}) == 1:
        return viable[0], similarity_score(report_row.cliente, viable[0].cliente_nombre), False
    if len(viable) == 1:
        return viable[0], similarity_score(report_row.cliente, viable[0].cliente_nombre), False

    best_candidate = None
    best_score = -1.0
    for candidate in viable:
        score = similarity_score(report_row.cliente, candidate.cliente_nombre)
        if score > best_score or (
            score == best_score and best_candidate is not None and candidate.current_rank < best_candidate.current_rank
        ):
            best_candidate = candidate
            best_score = score

    if best_candidate is not None and best_score >= 0.45:
        return best_candidate, max(best_score, 0.0), False

    return viable[0], max(best_score, 0.0), True


def build_updates(
    report_rows: Iterable[ReportRow],
    equipment_by_serial: dict[str, list[EquipmentRow]],
) -> tuple[dict[int, tuple[str, str]], dict[str, int]]:
    client_contact_signals: dict[int, dict[str, SignalAggregate]] = defaultdict(lambda: defaultdict(SignalAggregate))
    client_phone_signals: dict[int, dict[str, SignalAggregate]] = defaultdict(lambda: defaultdict(SignalAggregate))
    client_pair_signals: dict[int, dict[tuple[str, str], SignalAggregate]] = defaultdict(lambda: defaultdict(SignalAggregate))

    stats = Counter()
    unmatched_serials: set[str] = set()

    for report_row in report_rows:
        stats["report_rows"] += 1
        candidates = equipment_by_serial.get(report_row.serial_norm)
        if not candidates:
            unmatched_serials.add(report_row.serial_norm)
            continue

        if len(candidates) > 1:
            stats["duplicate_serial_rows"] += 1

        equipment, match_score, used_fallback = choose_equipment(report_row, candidates)
        if equipment is None or equipment.cliente_id is None:
            stats["rows_without_cliente_id"] += 1
            continue

        if len(candidates) > 1 and used_fallback:
            stats["low_confidence_duplicate_rows"] += 1
        if len(candidates) > 1 and not used_fallback:
            stats["resolved_duplicate_rows"] += 1

        base_weight = 1.0
        if match_score >= 0.92:
            base_weight += 1.5
        elif match_score >= 0.75:
            base_weight += 1.0
        elif match_score >= 0.45:
            base_weight += 0.5

        if report_row.contacto_key:
            client_contact_signals[equipment.cliente_id][report_row.contacto_key].add(
                report_row.contacto,
                base_weight + (0.75 if len(report_row.contacto_key) >= 10 else 0.25),
                report_row.fecha_servicio,
            )
            stats["rows_with_contact_signal"] += 1

        if report_row.telefono_key:
            phone_bonus = 1.0 if len(report_row.telefono_key) == 10 else 0.5
            client_phone_signals[equipment.cliente_id][report_row.telefono_key].add(
                report_row.telefono,
                base_weight + phone_bonus,
                report_row.fecha_servicio,
            )
            stats["rows_with_phone_signal"] += 1

        if report_row.contacto_key and report_row.telefono_key:
            client_pair_signals[equipment.cliente_id][(report_row.contacto_key, report_row.telefono_key)].add(
                f"{report_row.contacto}|||{report_row.telefono}",
                base_weight + 2.0,
                report_row.fecha_servicio,
            )
            stats["rows_with_pair_signal"] += 1

        stats["matched_rows"] += 1

    updates: dict[int, tuple[str, str]] = {}
    for cliente_id in sorted(set(client_contact_signals) | set(client_phone_signals) | set(client_pair_signals)):
        contact = ""
        phone = ""

        pairs = client_pair_signals.get(cliente_id, {})
        contacts = client_contact_signals.get(cliente_id, {})
        phones = client_phone_signals.get(cliente_id, {})

        if pairs:
            best_pair_key, best_pair_signal = sorted(pairs.items(), key=signal_sort_key, reverse=True)[0]
            pair_count = best_pair_signal.count
            pair_weight = best_pair_signal.weight
            if pair_count >= 2 or pair_weight >= 4.5:
                contact_key, phone_key = best_pair_key
                if contact_key in contacts:
                    contact = choose_variant(contacts[contact_key])
                if phone_key in phones:
                    phone = choose_variant(phones[phone_key])
                if contact or phone:
                    stats["clients_selected_from_pair"] += 1

        if not contact and contacts:
            contact = choose_variant(sorted(contacts.items(), key=signal_sort_key, reverse=True)[0][1])
        if not phone and phones:
            phone = choose_variant(sorted(phones.items(), key=signal_sort_key, reverse=True)[0][1])

        if contact or phone:
            updates[cliente_id] = (contact, phone)

    stats["unique_unmatched_serials"] = len(unmatched_serials)
    stats["clients_with_updates"] = len(updates)
    return updates, dict(stats)


def apply_updates(updates: dict[int, tuple[str, str]]) -> None:
    if not updates:
        return

    values = ",\n".join(
        f"({cliente_id}, {sql_literal(contact)}, {sql_literal(phone)})"
        for cliente_id, (contact, phone) in sorted(updates.items())
    )
    sql = f"""
    WITH updates(cliente_id, persona_contacto, telefono) AS (
      VALUES
      {values}
    )
    UPDATE public.clientes AS clientes
    SET
      persona_contacto = COALESCE(NULLIF(updates.persona_contacto, ''), clientes.persona_contacto),
      telefono = COALESCE(NULLIF(updates.telefono, ''), clientes.telefono)
    FROM updates
    WHERE clientes.id = updates.cliente_id;
    """
    execute_sql(sql)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Importa persona_contacto y telefono a clientes a partir de reportes historicos."
    )
    parser.add_argument("workbook", type=Path, help="Ruta del archivo Excel consolidado de reportes.")
    parser.add_argument("--apply", action="store_true", help="Aplica los cambios en la base de datos local.")
    args = parser.parse_args()

    report_rows = load_report_rows(args.workbook)
    equipment_by_serial = load_equipment_rows()
    updates, stats = build_updates(report_rows, equipment_by_serial)

    print(f"Filas de reportes con serial y contacto/telefono util: {stats.get('report_rows', 0)}")
    print(f"Filas matcheadas a cliente_id: {stats.get('matched_rows', 0)}")
    print(f"Filas con contacto util: {stats.get('rows_with_contact_signal', 0)}")
    print(f"Filas con telefono util: {stats.get('rows_with_phone_signal', 0)}")
    print(f"Filas con par contacto+telefono: {stats.get('rows_with_pair_signal', 0)}")
    print(f"Resoluciones de serial duplicado con similitud aceptable: {stats.get('resolved_duplicate_rows', 0)}")
    print(f"Resoluciones de serial duplicado con baja similitud: {stats.get('low_confidence_duplicate_rows', 0)}")
    print(f"Series sin match en equipos: {stats.get('unique_unmatched_serials', 0)}")
    print(f"Clientes con actualizacion preparada: {stats.get('clients_with_updates', 0)}")
    print(f"Clientes resueltos por par dominante contacto+telefono: {stats.get('clients_selected_from_pair', 0)}")

    for cliente_id, (contact, phone) in list(sorted(updates.items()))[:10]:
        print(f"cliente_id={cliente_id}\tcontacto={contact}\ttelefono={phone}")

    if args.apply:
        apply_updates(updates)
        print(f"Actualizacion aplicada a {len(updates)} clientes.")


if __name__ == "__main__":
    main()
