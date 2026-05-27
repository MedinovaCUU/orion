#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import subprocess
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


DB_CONTAINER = "supabase_db_Biosystems_Project"
DB_NAME = "postgres"
DB_USER = "postgres"


STATE_ALIASES = {
    "CDMX": "CIUDADDEMEXICO",
    "CIUDADDEMEXICO": "CIUDADDEMEXICO",
    "EDOMEX": "ESTADODEMEXICO",
    "ESTADODEMEXICO": "ESTADODEMEXICO",
}


@dataclass(frozen=True)
class WorkbookRow:
    serial_raw: str
    serial_norm: str
    client_code: str
    client_name: str
    laboratory: str
    model: str
    state: str
    municipio: str
    institution: str
    preventive_month: str
    osmosis: str
    osmosis_type: str


@dataclass(frozen=True)
class ClientRow:
    id: int
    id_original: str
    razon_social: str


@dataclass(frozen=True)
class EquipmentRow:
    id: str
    numero_serie: str
    serial_norm: str
    cliente_id: int | None
    cliente_codigo: str
    cliente_nombre: str
    modelo: str
    estado: str
    municipio: str
    current_rank: int


@dataclass(frozen=True)
class PendingClientInsert:
    client_code: str
    client_name: str


@dataclass(frozen=True)
class PendingClientCodeFill:
    client_id: int
    client_code: str


@dataclass(frozen=True)
class EquipmentInsert:
    id: str
    numero_serie: str
    cliente_id: int | None
    modelo: str
    estado: str
    municipio: str


@dataclass(frozen=True)
class EquipmentUpdate:
    equipment_id: str
    set_clauses: tuple[str, ...]


@dataclass
class ImportPlan:
    client_inserts: list[PendingClientInsert]
    client_code_fills: list[PendingClientCodeFill]
    equipment_inserts: list[EquipmentInsert]
    equipment_updates: list[EquipmentUpdate]
    stats: Counter[str]
    conflict_samples: list[str]
    unresolved_client_samples: list[str]


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"-", "N/A", "NA", "None"}:
        return ""
    return text


def normalize_name(value: object) -> str:
    text = clean_text(value)
    text = "".join(
        character for character in unicodedata.normalize("NFD", text) if unicodedata.category(character) != "Mn"
    )
    return re.sub(r"[^A-Z0-9]+", "", text.upper())


def normalize_state(value: object) -> str:
    normalized = normalize_name(value)
    return STATE_ALIASES.get(normalized, normalized)


def normalize_serial(value: object) -> str:
    digits = re.sub(r"\D+", "", clean_text(value))
    digits = digits.lstrip("0")
    return digits or "0"


def pick_mode(values: Iterable[str]) -> str:
    filtered = [value for value in values if clean_text(value)]
    if not filtered:
        return ""
    counts = Counter(filtered)
    return sorted(counts.items(), key=lambda item: (-item[1], -len(item[0]), item[0]))[0][0]


def sql_literal(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


def run_psql(sql: str) -> str:
    result = subprocess.run(
        [
            "docker",
            "exec",
            DB_CONTAINER,
            "psql",
            "-U",
            DB_USER,
            "-d",
            DB_NAME,
            "-At",
            "-F",
            "\t",
            "-c",
            sql,
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    return result.stdout


def execute_sql(sql: str) -> None:
    subprocess.run(
        [
            "docker",
            "exec",
            "-i",
            DB_CONTAINER,
            "psql",
            "-U",
            DB_USER,
            "-d",
            DB_NAME,
            "-v",
            "ON_ERROR_STOP=1",
        ],
        check=True,
        input=sql,
        text=True,
    )


def find_information_sheet(workbook) -> str:
    for sheet_name in workbook.sheetnames:
        if normalize_name(sheet_name) == "INFORMACION":
            return sheet_name
    raise ValueError("No se encontro una hoja equivalente a INFORMACION en el workbook.")


def merge_rows(rows: list[WorkbookRow]) -> WorkbookRow:
    return WorkbookRow(
        serial_raw=pick_mode(row.serial_raw for row in rows),
        serial_norm=rows[0].serial_norm,
        client_code=pick_mode(row.client_code for row in rows),
        client_name=pick_mode(row.client_name for row in rows),
        laboratory=pick_mode(row.laboratory for row in rows),
        model=pick_mode(row.model for row in rows),
        state=pick_mode(row.state for row in rows),
        municipio=pick_mode(row.municipio for row in rows),
        institution=pick_mode(row.institution for row in rows),
        preventive_month=pick_mode(row.preventive_month for row in rows),
        osmosis=pick_mode(row.osmosis for row in rows),
        osmosis_type=pick_mode(row.osmosis_type for row in rows),
    )


def load_workbook_rows(workbook_path: Path) -> list[WorkbookRow]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_name = find_information_sheet(workbook)
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [clean_text(cell) for cell in rows[0]]
    required = {"Entidad", "Municipio", "# Cliente", "Cliente", "Laboratorio", "Serial", "Modelo"}
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(f"Faltan columnas requeridas en la hoja {sheet_name}: {', '.join(missing)}")

    grouped: dict[str, list[WorkbookRow]] = defaultdict(list)
    for row in rows[1:]:
        if not any(value is not None and clean_text(value) for value in row):
            continue

        record = {headers[index]: row[index] for index in range(len(headers))}
        serial_norm = normalize_serial(record.get("Serial"))
        serial_raw = clean_text(record.get("Serial"))
        if serial_norm == "0" or not serial_raw:
            continue

        grouped[serial_norm].append(
            WorkbookRow(
                serial_raw=serial_raw,
                serial_norm=serial_norm,
                client_code=clean_text(record.get("# Cliente")),
                client_name=clean_text(record.get("Cliente")),
                laboratory=clean_text(record.get("Laboratorio")),
                model=clean_text(record.get("Modelo")),
                state=clean_text(record.get("Entidad")),
                municipio=clean_text(record.get("Municipio")),
                institution=clean_text(record.get("Institucion")) or clean_text(record.get("Institución")),
                preventive_month=clean_text(record.get("Mtto Preventivo")),
                osmosis=clean_text(record.get("Osmosis")),
                osmosis_type=clean_text(record.get("Tipo de Osmosis")),
            )
        )

    return [merge_rows(grouped[key]) for key in sorted(grouped)]


def load_clients() -> tuple[dict[str, ClientRow], dict[str, list[ClientRow]]]:
    sql = """
    SELECT id, COALESCE(id_original, ''), COALESCE(razon_social, '')
    FROM public.clientes
    ORDER BY id;
    """
    rows = run_psql(sql).splitlines()
    by_code: dict[str, ClientRow] = {}
    by_name: dict[str, list[ClientRow]] = defaultdict(list)
    for row in rows:
        if not row.strip():
            continue
        client_id_text, id_original, razon_social = row.split("\t")
        client = ClientRow(id=int(client_id_text), id_original=id_original, razon_social=razon_social)
        if id_original:
            by_code[id_original] = client
        by_name[normalize_name(razon_social)].append(client)
    return by_code, by_name


def load_equipment_rows() -> dict[str, list[EquipmentRow]]:
    sql = """
    WITH ranked AS (
      SELECT
        e.id,
        e.numero_serie,
        e.cliente_id,
        COALESCE(c.id_original, '') AS cliente_codigo,
        COALESCE(c.razon_social, '') AS cliente_nombre,
        COALESCE(e.modelo, '') AS modelo,
        COALESCE(e.estado, '') AS estado,
        COALESCE(e.municipio, '') AS municipio,
        ROW_NUMBER() OVER (
          PARTITION BY e.numero_serie
          ORDER BY
            (e.fecha_fin IS NULL) DESC,
            e.fecha_fin DESC NULLS LAST,
            e.fecha_inicio DESC NULLS LAST,
            e.actualizado_en DESC NULLS LAST,
            e.creado_en DESC,
            e.id DESC
        ) AS rn
      FROM public.equipos AS e
      LEFT JOIN public.clientes AS c ON c.id = e.cliente_id
    )
    SELECT
      id,
      numero_serie,
      COALESCE(cliente_id::text, ''),
      cliente_codigo,
      cliente_nombre,
      modelo,
      estado,
      municipio,
      rn
    FROM ranked
    ORDER BY numero_serie, rn;
    """
    rows = run_psql(sql).splitlines()
    by_serial: dict[str, list[EquipmentRow]] = defaultdict(list)
    for row in rows:
        if not row.strip():
            continue
        (
            equipment_id,
            numero_serie,
            cliente_id_text,
            cliente_codigo,
            cliente_nombre,
            modelo,
            estado,
            municipio,
            current_rank_text,
        ) = row.split("\t")
        serial_norm = normalize_serial(numero_serie)
        by_serial[serial_norm].append(
            EquipmentRow(
                id=equipment_id,
                numero_serie=numero_serie,
                serial_norm=serial_norm,
                cliente_id=int(cliente_id_text) if cliente_id_text else None,
                cliente_codigo=cliente_codigo,
                cliente_nombre=cliente_nombre,
                modelo=modelo,
                estado=estado,
                municipio=municipio,
                current_rank=int(current_rank_text),
            )
        )
    return by_serial


def load_next_equipment_id() -> int:
    sql = "SELECT COALESCE(MAX(id::bigint), 0) FROM public.equipos WHERE id ~ '^[0-9]+$';"
    return int(run_psql(sql).strip() or "0") + 1


def resolve_client(
    workbook_row: WorkbookRow,
    clients_by_code: dict[str, ClientRow],
    clients_by_name: dict[str, list[ClientRow]],
) -> tuple[ClientRow | None, PendingClientInsert | None, PendingClientCodeFill | None, str | None]:
    normalized_name = normalize_name(workbook_row.client_name)

    if workbook_row.client_code and workbook_row.client_code in clients_by_code:
        return clients_by_code[workbook_row.client_code], None, None, None

    exact_name_matches = clients_by_name.get(normalized_name, [])

    if not workbook_row.client_code:
        if len(exact_name_matches) == 1:
            return exact_name_matches[0], None, None, None
        if workbook_row.client_name:
            return None, PendingClientInsert(client_code="", client_name=workbook_row.client_name), None, None
        return None, None, None, "missing_client_name"

    if len(exact_name_matches) == 1 and not exact_name_matches[0].id_original:
        return exact_name_matches[0], None, PendingClientCodeFill(exact_name_matches[0].id, workbook_row.client_code), None

    if workbook_row.client_name:
        return None, PendingClientInsert(workbook_row.client_code, workbook_row.client_name), None, None

    return None, None, None, "missing_client_name"


def choose_target_equipment(workbook_row: WorkbookRow, candidates: list[EquipmentRow], client_id: int | None) -> EquipmentRow:
    if client_id is not None:
        same_client = [candidate for candidate in candidates if candidate.cliente_id == client_id]
        if same_client:
            return min(same_client, key=lambda item: item.current_rank)

    if len(candidates) == 1:
        return candidates[0]

    return min(candidates, key=lambda item: item.current_rank)


def build_plan(
    workbook_rows: list[WorkbookRow],
    clients_by_code: dict[str, ClientRow],
    clients_by_name: dict[str, list[ClientRow]],
    equipment_by_serial: dict[str, list[EquipmentRow]],
    next_equipment_id: int,
) -> ImportPlan:
    client_inserts: dict[tuple[str, str], PendingClientInsert] = {}
    client_code_fills: dict[int, PendingClientCodeFill] = {}
    equipment_inserts: list[EquipmentInsert] = []
    equipment_updates: list[EquipmentUpdate] = []
    stats: Counter[str] = Counter()
    conflict_samples: list[str] = []
    unresolved_client_samples: list[str] = []

    pending_client_ids: dict[tuple[str, str], int] = {}
    pending_id = -1
    next_insert_id = next_equipment_id

    for workbook_row in workbook_rows:
        stats["workbook_rows"] += 1

        client, pending_insert, pending_code_fill, unresolved_reason = resolve_client(
            workbook_row,
            clients_by_code,
            clients_by_name,
        )

        resolved_client_id: int | None = client.id if client is not None else None
        if client is not None:
            stats["resolved_clients"] += 1

        if pending_code_fill is not None:
            existing_fill = client_code_fills.get(pending_code_fill.client_id)
            if existing_fill is None:
                client_code_fills[pending_code_fill.client_id] = pending_code_fill
                stats["client_code_fills"] += 1
            resolved_client_id = pending_code_fill.client_id

        if pending_insert is not None:
            key = (pending_insert.client_code, normalize_name(pending_insert.client_name))
            if key not in client_inserts:
                client_inserts[key] = pending_insert
                pending_client_ids[key] = pending_id
                pending_id -= 1
                stats["client_inserts"] += 1
            resolved_client_id = pending_client_ids[key]

        if unresolved_reason is not None:
            stats["unresolved_clients"] += 1
            if len(unresolved_client_samples) < 12:
                unresolved_client_samples.append(
                    f"serial={workbook_row.serial_raw} client_code={workbook_row.client_code or '-'} "
                    f"client_name={workbook_row.client_name or '-'} reason={unresolved_reason}"
                )

        candidates = equipment_by_serial.get(workbook_row.serial_norm, [])
        if not candidates:
            equipment_inserts.append(
                EquipmentInsert(
                    id=str(next_insert_id),
                    numero_serie=workbook_row.serial_raw,
                    cliente_id=resolved_client_id if resolved_client_id and resolved_client_id > 0 else None,
                    modelo=workbook_row.model,
                    estado=workbook_row.state,
                    municipio=workbook_row.municipio,
                )
            )
            next_insert_id += 1
            stats["equipment_inserts"] += 1
            continue

        target = choose_target_equipment(workbook_row, candidates, resolved_client_id if resolved_client_id and resolved_client_id > 0 else None)
        set_clauses: list[str] = []

        if resolved_client_id and resolved_client_id > 0:
            if target.cliente_id is None:
                set_clauses.append(f"cliente_id = {resolved_client_id}")
                stats["equipment_client_fills"] += 1
            elif target.cliente_id != resolved_client_id:
                stats["client_conflicts"] += 1
                if len(conflict_samples) < 20:
                    conflict_samples.append(
                        f"client_conflict serial={workbook_row.serial_raw} equipment_id={target.id} "
                        f"db_client={target.cliente_codigo or '-'}:{target.cliente_nombre or '-'} "
                        f"excel_client={workbook_row.client_code or '-'}:{workbook_row.client_name or '-'}"
                    )

        if workbook_row.model:
            if not clean_text(target.modelo):
                set_clauses.append(f"modelo = {sql_literal(workbook_row.model)}")
                stats["equipment_model_fills"] += 1
            elif normalize_name(target.modelo) != normalize_name(workbook_row.model):
                stats["model_conflicts"] += 1
                if len(conflict_samples) < 20:
                    conflict_samples.append(
                        f"model_conflict serial={workbook_row.serial_raw} equipment_id={target.id} "
                        f"db_model={target.modelo or '-'} excel_model={workbook_row.model or '-'}"
                    )

        if workbook_row.state:
            if not clean_text(target.estado):
                set_clauses.append(f"estado = {sql_literal(workbook_row.state)}")
                stats["equipment_state_fills"] += 1
            elif normalize_state(target.estado) != normalize_state(workbook_row.state):
                stats["state_conflicts"] += 1
                if len(conflict_samples) < 20:
                    conflict_samples.append(
                        f"state_conflict serial={workbook_row.serial_raw} equipment_id={target.id} "
                        f"db_state={target.estado or '-'} excel_state={workbook_row.state or '-'}"
                    )

        if workbook_row.municipio:
            if not clean_text(target.municipio):
                set_clauses.append(f"municipio = {sql_literal(workbook_row.municipio)}")
                stats["equipment_municipio_fills"] += 1
            elif normalize_name(target.municipio) != normalize_name(workbook_row.municipio):
                stats["municipio_conflicts"] += 1
                if len(conflict_samples) < 20:
                    conflict_samples.append(
                        f"municipio_conflict serial={workbook_row.serial_raw} equipment_id={target.id} "
                        f"db_municipio={target.municipio or '-'} excel_municipio={workbook_row.municipio or '-'}"
                    )

        if set_clauses:
            if not any(clause.startswith("actualizado_en = ") for clause in set_clauses):
                set_clauses.append("actualizado_en = NOW()")
            equipment_updates.append(EquipmentUpdate(target.id, tuple(set_clauses)))
            stats["equipment_updates"] += 1
        else:
            stats["equipment_noops"] += 1

    return ImportPlan(
        client_inserts=sorted(client_inserts.values(), key=lambda item: (item.client_code, normalize_name(item.client_name))),
        client_code_fills=sorted(client_code_fills.values(), key=lambda item: item.client_id),
        equipment_inserts=equipment_inserts,
        equipment_updates=equipment_updates,
        stats=stats,
        conflict_samples=conflict_samples,
        unresolved_client_samples=unresolved_client_samples,
    )


def build_client_sql(plan: ImportPlan) -> str:
    statements: list[str] = ["BEGIN;"]

    for item in plan.client_inserts:
        statements.append(
            "INSERT INTO public.clientes (id_original, razon_social) "
            f"VALUES ({sql_literal(item.client_code or None)}, {sql_literal(item.client_name)}) "
            "RETURNING id;"
        )

    for item in plan.client_code_fills:
        statements.append(
            f"UPDATE public.clientes SET id_original = {sql_literal(item.client_code)} "
            f"WHERE id = {item.client_id} AND COALESCE(id_original, '') = '';"
        )

    statements.append("COMMIT;")
    return "\n".join(statements)


def build_equipment_sql(plan: ImportPlan) -> str:
    statements: list[str] = ["BEGIN;"]

    for item in plan.equipment_inserts:
        statements.append(
            "INSERT INTO public.equipos (id, numero_serie, cliente_id, modelo, estado, municipio) "
            "VALUES ("
            f"{sql_literal(item.id)}, "
            f"{sql_literal(item.numero_serie)}, "
            f"{str(item.cliente_id) if item.cliente_id is not None else 'NULL'}, "
            f"{sql_literal(item.modelo or None)}, "
            f"{sql_literal(item.estado or None)}, "
            f"{sql_literal(item.municipio or None)}"
            ");"
        )

    for item in plan.equipment_updates:
        statements.append(
            f"UPDATE public.equipos SET {', '.join(item.set_clauses)} WHERE id = {sql_literal(item.equipment_id)};"
        )

    statements.append("COMMIT;")
    return "\n".join(statements)


def print_summary(plan: ImportPlan) -> None:
    print(f"Rows in workbook: {plan.stats.get('workbook_rows', 0)}")
    print(f"Resolved clients: {plan.stats.get('resolved_clients', 0)}")
    print(f"Client inserts planned: {len(plan.client_inserts)}")
    print(f"Client id_original fills planned: {len(plan.client_code_fills)}")
    print(f"Equipment inserts planned: {len(plan.equipment_inserts)}")
    print(f"Equipment updates planned: {len(plan.equipment_updates)}")
    print(f"Equipment no-op rows: {plan.stats.get('equipment_noops', 0)}")
    print(f"Safe fills -> client_id: {plan.stats.get('equipment_client_fills', 0)}")
    print(f"Safe fills -> modelo: {plan.stats.get('equipment_model_fills', 0)}")
    print(f"Safe fills -> estado: {plan.stats.get('equipment_state_fills', 0)}")
    print(f"Safe fills -> municipio: {plan.stats.get('equipment_municipio_fills', 0)}")
    print(f"Conflicts skipped -> client_id: {plan.stats.get('client_conflicts', 0)}")
    print(f"Conflicts skipped -> modelo: {plan.stats.get('model_conflicts', 0)}")
    print(f"Conflicts skipped -> estado: {plan.stats.get('state_conflicts', 0)}")
    print(f"Conflicts skipped -> municipio: {plan.stats.get('municipio_conflicts', 0)}")
    print(f"Unresolved clients: {plan.stats.get('unresolved_clients', 0)}")
    print(
        "Unmapped workbook columns retained outside this import: "
        "Laboratorio, Institucion, Mtto Preventivo, Osmosis, Tipo de Osmosis, 24 hrs prendido, "
        "lamparas anuales, Ano de fabricacion."
    )

    if plan.equipment_inserts:
        print("Sample new equipment rows:")
        for item in plan.equipment_inserts[:10]:
            print(
                f"  serial={item.numero_serie} id={item.id} client_id={item.cliente_id or 'NULL'} "
                f"modelo={item.modelo or '-'} estado={item.estado or '-'} municipio={item.municipio or '-'}"
            )

    if plan.unresolved_client_samples:
        print("Unresolved client samples:")
        for sample in plan.unresolved_client_samples:
            print(f"  {sample}")

    if plan.conflict_samples:
        print("Conflict samples:")
        for sample in plan.conflict_samples[:20]:
            print(f"  {sample}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Importa la hoja INFORMACION de un workbook de gobierno hacia clientes/equipos "
            "sin sobrescribir conflictos: solo inserta faltantes y completa campos vacios."
        )
    )
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("--apply", action="store_true", help="Aplica los cambios a la base local. Sin esto, solo muestra el resumen.")
    args = parser.parse_args()

    workbook_rows = load_workbook_rows(args.xlsx_path)
    clients_by_code, clients_by_name = load_clients()
    equipment_by_serial = load_equipment_rows()
    next_equipment_id = load_next_equipment_id()

    plan = build_plan(workbook_rows, clients_by_code, clients_by_name, equipment_by_serial, next_equipment_id)
    print_summary(plan)

    if not args.apply:
        return

    if plan.client_inserts or plan.client_code_fills:
        execute_sql(build_client_sql(plan))
        clients_by_code, clients_by_name = load_clients()

    equipment_by_serial = load_equipment_rows()
    next_equipment_id = load_next_equipment_id()
    plan = build_plan(workbook_rows, clients_by_code, clients_by_name, equipment_by_serial, next_equipment_id)

    if plan.unresolved_client_samples:
        raise RuntimeError("Hay clientes sin resolver despues de refrescar el catalogo. Revisar dry-run antes de aplicar.")

    if plan.equipment_inserts or plan.equipment_updates:
        execute_sql(build_equipment_sql(plan))

    clients_by_code, clients_by_name = load_clients()
    equipment_by_serial = load_equipment_rows()
    next_equipment_id = load_next_equipment_id()
    final_plan = build_plan(workbook_rows, clients_by_code, clients_by_name, equipment_by_serial, next_equipment_id)

    print("---")
    print("Applied changes.")
    print_summary(final_plan)


if __name__ == "__main__":
    main()
