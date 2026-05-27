#!/usr/bin/env python3
from __future__ import annotations

import argparse
import difflib
import re
import subprocess
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
import unicodedata

from openpyxl import load_workbook


DB_CONTAINER = "supabase_db_Biosystems_Project"
DB_NAME = "postgres"
DB_USER = "postgres"


@dataclass
class WorkbookRow:
    serial_norm: str
    cliente: str
    direccion: str
    software: str
    firmware: str
    contacto: str
    telefono: str


@dataclass
class EquipmentRow:
    id: str
    numero_serie: str
    serial_norm: str
    cliente_id: int | None
    cliente_nombre: str
    current_rank: int


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"-", "N/A", "NA", "None"}:
        return ""
    return text


def normalize_serial(value: object) -> str:
    digits = re.sub(r"\D+", "", clean_text(value))
    digits = digits.lstrip("0")
    return digits or "0"


def normalize_name(value: object) -> str:
    text = clean_text(value)
    text = "".join(
        character for character in unicodedata.normalize("NFD", text) if unicodedata.category(character) != "Mn"
    )
    return re.sub(r"[^A-Z0-9]+", "", text.upper())


def pick_mode(values: Iterable[str]) -> str:
    filtered = [value for value in values if clean_text(value)]
    if not filtered:
        return ""

    counts = Counter(filtered)
    return sorted(counts.items(), key=lambda item: (-item[1], -len(item[0]), item[0]))[0][0]


def sql_literal(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


def run_psql(sql: str) -> str:
    result = subprocess.run(
        [
            "docker",
            "exec",
            DB_CONTAINER,
            "psql",
            "-U",
            DB_USER,
            "-d",
            DB_NAME,
            "-At",
            "-F",
            "\t",
            "-c",
            sql,
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    return result.stdout


def execute_sql(sql: str) -> None:
    subprocess.run(
        [
            "docker",
            "exec",
            "-i",
            DB_CONTAINER,
            "psql",
            "-U",
            DB_USER,
            "-d",
            DB_NAME,
            "-v",
            "ON_ERROR_STOP=1",
        ],
        check=True,
        input=sql,
        text=True,
    )


def similarity_score(left: str, right: str) -> float:
    left_norm = normalize_name(left)
    right_norm = normalize_name(right)
    if not left_norm or not right_norm:
        return 0.0
    if left_norm == right_norm:
        return 1.0
    if left_norm in right_norm or right_norm in left_norm:
        return 0.92
    return difflib.SequenceMatcher(None, left_norm, right_norm).ratio()


def load_workbook_rows(workbook_path: Path) -> dict[str, list[WorkbookRow]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [clean_text(cell) for cell in rows[0]]

    by_serial: dict[str, list[WorkbookRow]] = defaultdict(list)
    for row in rows[1:]:
        if not any(value is not None and clean_text(value) for value in row):
            continue
        record = {headers[index]: row[index] for index in range(len(headers))}
        serial_norm = normalize_serial(record.get("Numero de Serie"))
        if not serial_norm or serial_norm == "0":
            continue
        by_serial[serial_norm].append(
            WorkbookRow(
                serial_norm=serial_norm,
                cliente=clean_text(record.get("Cliente")),
                direccion=clean_text(record.get("Direccion_Completa_Maps")),
                software=clean_text(record.get("Software Usuario")),
                firmware=clean_text(record.get("Firmware")),
                contacto=clean_text(record.get("Persona de contacto")),
                telefono=clean_text(record.get("Telefono")),
            )
        )

    return by_serial


def load_equipment_rows() -> dict[str, list[EquipmentRow]]:
    sql = """
    WITH ranked AS (
      SELECT
        equipos.id,
        equipos.numero_serie,
        equipos.cliente_id,
        COALESCE(clientes.razon_social, '') AS cliente_nombre,
        ROW_NUMBER() OVER (
          PARTITION BY equipos.numero_serie
          ORDER BY
            (equipos.fecha_fin IS NULL) DESC,
            equipos.fecha_fin DESC NULLS LAST,
            equipos.fecha_inicio DESC NULLS LAST,
            equipos.creado_en DESC,
            equipos.id DESC
        ) AS rn
      FROM public.equipos
      LEFT JOIN public.clientes ON clientes.id = equipos.cliente_id
    )
    SELECT id, numero_serie, COALESCE(cliente_id::text, ''), cliente_nombre, rn
    FROM ranked
    ORDER BY numero_serie, rn;
    """
    rows = run_psql(sql).splitlines()
    equipment_by_serial: dict[str, list[EquipmentRow]] = defaultdict(list)

    for row in rows:
        if not row.strip():
            continue
        equipment_id, numero_serie, cliente_id_text, cliente_nombre, current_rank = row.split("\t")
        serial_norm = normalize_serial(numero_serie)
        equipment_by_serial[serial_norm].append(
            EquipmentRow(
                id=equipment_id,
                numero_serie=numero_serie,
                serial_norm=serial_norm,
                cliente_id=int(cliente_id_text) if cliente_id_text else None,
                cliente_nombre=cliente_nombre,
                current_rank=int(current_rank),
            )
        )

    return equipment_by_serial


def pick_mode_row(rows: list[WorkbookRow]) -> WorkbookRow:
    return WorkbookRow(
        serial_norm=rows[0].serial_norm,
        cliente=pick_mode(row.cliente for row in rows),
        direccion=pick_mode(row.direccion for row in rows),
        software=pick_mode(row.software for row in rows),
        firmware=pick_mode(row.firmware for row in rows),
        contacto=pick_mode(row.contacto for row in rows),
        telefono=pick_mode(row.telefono for row in rows),
    )


def choose_workbook_row(candidate_rows: list[WorkbookRow], equipment_row: EquipmentRow) -> WorkbookRow:
    if len(candidate_rows) == 1:
        return candidate_rows[0]

    best_row = None
    best_score = -1.0
    for row in candidate_rows:
        score = similarity_score(row.cliente, equipment_row.cliente_nombre)
        if score > best_score:
            best_row = row
            best_score = score

    if best_row and best_score >= 0.28:
        return best_row

    return pick_mode_row(candidate_rows)


def build_update_plan(
    workbook_rows: dict[str, list[WorkbookRow]],
    equipment_rows: dict[str, list[EquipmentRow]],
) -> tuple[list[str], dict[int, dict[str, list[str]]], set[str], int]:
    equipment_updates: list[str] = []
    client_signals: dict[int, dict[str, list[str]]] = defaultdict(lambda: {"contactos": [], "telefonos": []})
    unmatched_serials: set[str] = set()
    duplicate_address_matches = 0

    for serial_norm, workbook_candidates in workbook_rows.items():
        equipment_candidates = equipment_rows.get(serial_norm)
        if not equipment_candidates:
            unmatched_serials.add(serial_norm)
            continue

        for equipment in equipment_candidates:
            chosen_row = choose_workbook_row(workbook_candidates, equipment)
            set_clauses: list[str] = []
            if chosen_row.software:
                set_clauses.append(f"\"Software\" = {sql_literal(chosen_row.software)}")
            if chosen_row.firmware:
                set_clauses.append(f"\"Firmware\" = {sql_literal(chosen_row.firmware)}")
            if chosen_row.direccion:
                set_clauses.append(f"direccion = {sql_literal(chosen_row.direccion)}")
                if len(equipment_candidates) > 1:
                    duplicate_address_matches += 1

            if set_clauses:
                equipment_updates.append(
                    f"UPDATE public.equipos SET {', '.join(set_clauses)} WHERE id = {sql_literal(equipment.id)};"
                )

            if equipment.cliente_id is not None:
                if chosen_row.contacto:
                    client_signals[equipment.cliente_id]["contactos"].append(chosen_row.contacto)
                if chosen_row.telefono:
                    client_signals[equipment.cliente_id]["telefonos"].append(chosen_row.telefono)

    return equipment_updates, client_signals, unmatched_serials, duplicate_address_matches


def build_client_updates(client_signals: dict[int, dict[str, list[str]]]) -> list[str]:
    updates: list[str] = []
    for client_id, signals in client_signals.items():
        contacto = pick_mode(signals["contactos"])
        telefono = pick_mode(signals["telefonos"])

        set_clauses: list[str] = []
        if contacto:
            set_clauses.append(f"persona_contacto = {sql_literal(contacto)}")
        if telefono:
            set_clauses.append(f"telefono = {sql_literal(telefono)}")

        if set_clauses:
            updates.append(f"UPDATE public.clientes SET {', '.join(set_clauses)} WHERE id = {client_id};")
    return updates


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa datos de contacto/software/firmware desde la base maestra de equipos.")
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("--apply", action="store_true", help="Aplica los cambios a la base local. Sin esto, solo muestra el resumen.")
    args = parser.parse_args()

    workbook_rows = load_workbook_rows(args.xlsx_path)
    equipment_rows = load_equipment_rows()
    equipment_updates, client_signals, unmatched_serials, duplicate_address_matches = build_update_plan(
        workbook_rows,
        equipment_rows,
    )
    client_updates = build_client_updates(client_signals)

    print(f"Series unicas en Excel: {len(workbook_rows)}")
    print(f"Filas de equipos a actualizar: {len(equipment_updates)}")
    print(f"Clientes a enriquecer via cliente_id de equipos: {len(client_updates)}")
    print(f"Direcciones asignadas en equipos duplicados por similitud de cliente: {duplicate_address_matches}")
    print(f"Series sin match en equipos: {len(unmatched_serials)}")
    if unmatched_serials:
        print("Muestra sin match:", ", ".join(sorted(unmatched_serials)[:20]))

    if not args.apply:
        return

    sql = "\n".join(
        [
            "BEGIN;",
            *equipment_updates,
            *client_updates,
            "COMMIT;",
        ]
    )
    execute_sql(sql)
    print("Cambios aplicados en la base local.")


if __name__ == "__main__":
    main()
