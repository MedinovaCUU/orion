#!/usr/bin/env python3
from __future__ import annotations

import argparse
import difflib
import re
import subprocess
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


DB_CONTAINER = "supabase_db_Biosystems_Project"
DB_NAME = "postgres"
DB_USER = "postgres"


@dataclass
class WorkbookRow:
    serial_norm: str
    cliente: str
    direccion: str
    codigo_postal: str
    ciudad: str
    estado: str


@dataclass
class EquipmentRow:
    id: str
    numero_serie: str
    serial_norm: str
    cliente_nombre: str
    current_rank: int


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"-", "N/A", "NA", "None", "0000-00-00"}:
        return ""
    return text


def normalize_serial(value: object) -> str:
    digits = re.sub(r"\D+", "", clean_text(value))
    digits = digits.lstrip("0")
    return digits or "0"


def normalize_name(value: object) -> str:
    text = clean_text(value)
    text = "".join(
        character for character in unicodedata.normalize("NFD", text) if unicodedata.category(character) != "Mn"
    )
    return re.sub(r"[^A-Z0-9]+", "", text.upper())


def similarity_score(left: str, right: str) -> float:
    left_norm = normalize_name(left)
    right_norm = normalize_name(right)
    if not left_norm or not right_norm:
        return 0.0
    if left_norm == right_norm:
        return 1.0
    if left_norm in right_norm or right_norm in left_norm:
        return 0.92
    return difflib.SequenceMatcher(None, left_norm, right_norm).ratio()


def sql_literal(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


def run_psql(sql: str) -> str:
    result = subprocess.run(
        [
            "docker",
            "exec",
            DB_CONTAINER,
            "psql",
            "-U",
            DB_USER,
            "-d",
            DB_NAME,
            "-At",
            "-F",
            "\t",
            "-c",
            sql,
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    return result.stdout


def execute_sql(sql: str) -> None:
    subprocess.run(
        [
            "docker",
            "exec",
            "-i",
            DB_CONTAINER,
            "psql",
            "-U",
            DB_USER,
            "-d",
            DB_NAME,
            "-v",
            "ON_ERROR_STOP=1",
        ],
        check=True,
        input=sql,
        text=True,
    )


def load_workbook_rows(workbook_path: Path) -> dict[str, WorkbookRow]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [clean_text(cell) for cell in rows[0]]

    workbook_rows: dict[str, WorkbookRow] = {}
    for row in rows[1:]:
        if not any(value is not None and clean_text(value) for value in row):
            continue

        record = {headers[index]: row[index] for index in range(len(headers))}
        serial_norm = normalize_serial(record.get("Numero de Serie"))
        if not serial_norm or serial_norm == "0":
            continue

        workbook_rows[serial_norm] = WorkbookRow(
            serial_norm=serial_norm,
            cliente=clean_text(record.get("Cliente definitivo")),
            direccion=clean_text(record.get("Direccion definitiva")),
            codigo_postal=clean_text(record.get("Codigo Postal")),
            ciudad=clean_text(record.get("Ciudad")),
            estado=clean_text(record.get("Estado")),
        )

    return workbook_rows


def load_equipment_rows() -> dict[str, list[EquipmentRow]]:
    sql = """
    WITH ranked AS (
      SELECT
        equipos.id,
        equipos.numero_serie,
        COALESCE(clientes.razon_social, '') AS cliente_nombre,
        ROW_NUMBER() OVER (
          PARTITION BY equipos.numero_serie
          ORDER BY
            (equipos.fecha_fin IS NULL) DESC,
            equipos.fecha_fin DESC NULLS LAST,
            equipos.fecha_inicio DESC NULLS LAST,
            equipos.creado_en DESC,
            equipos.id DESC
        ) AS rn
      FROM public.equipos
      LEFT JOIN public.clientes ON clientes.id = equipos.cliente_id
    )
    SELECT id, numero_serie, cliente_nombre, rn
    FROM ranked
    ORDER BY numero_serie, rn;
    """
    rows = run_psql(sql).splitlines()
    equipment_by_serial: dict[str, list[EquipmentRow]] = defaultdict(list)

    for row in rows:
        if not row.strip():
            continue
        equipment_id, numero_serie, cliente_nombre, current_rank = row.split("\t")
        serial_norm = normalize_serial(numero_serie)
        equipment_by_serial[serial_norm].append(
            EquipmentRow(
                id=equipment_id,
                numero_serie=numero_serie,
                serial_norm=serial_norm,
                cliente_nombre=cliente_nombre,
                current_rank=int(current_rank),
            )
        )

    return equipment_by_serial


def choose_target_equipment(workbook_row: WorkbookRow, candidates: list[EquipmentRow]) -> tuple[EquipmentRow, float]:
    if len(candidates) == 1:
        return candidates[0], 1.0

    best_candidate = None
    best_score = -1.0
    for candidate in candidates:
        score = similarity_score(workbook_row.cliente, candidate.cliente_nombre)
        if score > best_score:
            best_candidate = candidate
            best_score = score

    if best_candidate is not None and best_score >= 0.28:
        return best_candidate, best_score

    fallback = min(candidates, key=lambda item: item.current_rank)
    return fallback, best_score


def build_update_plan(
    workbook_rows: dict[str, WorkbookRow],
    equipment_rows: dict[str, list[EquipmentRow]],
) -> tuple[list[str], int, int, set[str]]:
    updates: list[str] = []
    duplicated_matches = 0
    low_confidence_matches = 0
    unmatched_serials: set[str] = set()

    for serial_norm, workbook_row in workbook_rows.items():
        candidates = equipment_rows.get(serial_norm)
        if not candidates:
            unmatched_serials.add(serial_norm)
            continue

        target, score = choose_target_equipment(workbook_row, candidates)
        if len(candidates) > 1:
            duplicated_matches += 1
            if score < 0.28:
                low_confidence_matches += 1

        set_clauses: list[str] = []
        if workbook_row.direccion:
            set_clauses.append(f"direccion = {sql_literal(workbook_row.direccion)}")
        if workbook_row.ciudad:
            set_clauses.append(f"ciudad = {sql_literal(workbook_row.ciudad)}")
        if workbook_row.estado:
            set_clauses.append(f"estado = {sql_literal(workbook_row.estado)}")

        if not set_clauses:
            continue

        updates.append(f"UPDATE public.equipos SET {', '.join(set_clauses)} WHERE id = {sql_literal(target.id)};")

    return updates, duplicated_matches, low_confidence_matches, unmatched_serials


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa direccion, ciudad y estado de la base definitiva de equipos.")
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("--apply", action="store_true", help="Aplica los cambios a la base local. Sin esto, solo muestra el resumen.")
    args = parser.parse_args()

    workbook_rows = load_workbook_rows(args.xlsx_path)
    equipment_rows = load_equipment_rows()
    updates, duplicated_matches, low_confidence_matches, unmatched_serials = build_update_plan(workbook_rows, equipment_rows)

    print(f"Series unicas en Excel: {len(workbook_rows)}")
    print(f"Filas de equipos a actualizar con ciudad/estado: {len(updates)}")
    print(f"Series duplicadas resueltas por similitud de cliente: {duplicated_matches}")
    print(f"Duplicados con similitud baja y fallback al equipo vigente: {low_confidence_matches}")
    print(f"Series sin match en equipos: {len(unmatched_serials)}")
    if unmatched_serials:
        print("Muestra sin match:", ", ".join(sorted(unmatched_serials)[:20]))

    if not args.apply:
        return

    sql = "\n".join(["BEGIN;", *updates, "COMMIT;"])
    execute_sql(sql)
    print("Cambios aplicados en la base local.")


if __name__ == "__main__":
    main()
